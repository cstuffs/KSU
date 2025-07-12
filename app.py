from flask import Flask, render_template, request, redirect, url_for, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from collections import OrderedDict
from io import BytesIO
from openpyxl import Workbook
from sqlalchemy import text 
from sqlalchemy.orm import joinedload
from extensions import db
from models import db, Team, User, Order, OrderItem
from models import MenuGroup, MenuItem, MenuOption
from datetime import datetime, timedelta
import json
import os
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from apscheduler.schedulers.background import BackgroundScheduler
from tasks import email_all_orders

CDT = ZoneInfo("America/Chicago")

app = Flask(__name__)
app.secret_key = 'your-secret-key'

# Set DB URI
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get("DATABASE_URL", "sqlite:///local_fallback.db")
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize (but don’t define models here)
db.init_app(app)

# Setup LoginManager
login_manager = LoginManager()
login_manager.init_app(app)

# === Login Loader ===
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# === Create Tables Automatically ===
db_created = False

def get_week_number(date):
    """Returns week number where Week 1 is the week containing July 1st."""
    july_first = datetime(date.year, 7, 1).date()

    # Start weeks on Sunday
    if july_first.weekday() != 6:
        july_first -= timedelta(days=july_first.weekday() + 1)

    delta = date - july_first
    return (delta.days // 7) + 1

# === Example Home Route ===
@app.route('/')
def home():
    return "Welcome to the Team Ordering App!"

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        raw_team = request.form.get('team_name', '').strip()
        raw_name = request.form.get('member_name', '').strip()

        team_input = raw_team.lower()
        member_input = raw_name.lower()

        # Get matching team
        team = Team.query.filter(db.func.lower(Team.name) == team_input).first()
        if not team:
            return f"Team '{raw_team}' not found.", 403

        # Get matching user in that team
        user = User.query.filter(
            db.func.lower(User.name) == member_input,
            User.team_id == team.id
        ).first()

        if not user:
            return f"User '{raw_name}' not found on team '{raw_team}'.", 403

        # ❌ Block deactivated users
        if not user.is_enabled:
            return f"User '{raw_name}' is deactivated and cannot log in.", 403

        # ✅ Log the user in
        login_user(user)
        session['team'] = team.name
        session['member_name'] = user.name

        # ✅ Admin rules for KSU Football
        if team.name == "KSU Football":
            if user.name == "Scott Trausch":
                session['admin_as_football'] = False  # Full admin
            else:
                session['admin_as_football'] = True  # Limited admin
            return redirect(url_for('admin_dashboard'))

        # ✅ Standard user login
        session['admin_as_football'] = False
        return redirect(url_for('submit_order'))

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.clear()
    return redirect(url_for('login'))

@app.route('/order', methods=['GET'])
@login_required
def submit_order():
    # Redirect full admin to dashboard
    if current_user.name == 'admin' and not session.get('admin_as_football'):
        return redirect(url_for('admin_dashboard'))

    if request.args.get("new") == "1":
        session.pop("last_order_form", None)

    form_data = session.get('last_order_form', {})

    # ✅ Get team from database
    team = current_user.team
    if not team:
        return "Team not found for current user.", 400

    team_budget = team.budget
    remaining_budget = team.remaining_budget

    # ✅ Week range string
    today = datetime.now(CDT)
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)
    week_range_str = f"{start_of_week.strftime('%-m/%-d/%y')} - {end_of_week.strftime('%-m/%-d/%y')}"

    # ✅ Load structured menu
    from models import MenuGroup, MenuItem, MenuOption

    # Load menu from DB
    grouped_menu = OrderedDict()
    groups = MenuGroup.query.order_by(MenuGroup.id).all()
    for group in groups:
        group_data = OrderedDict()
        for item in group.items:
            options = [{
                "name": opt.name,
                "price": opt.price,
                "slug": (item.name + '_' + opt.name).replace(' ', '_').replace('(', '').replace(')', '').replace('/', '').replace('&', '').replace(',', '')
            } for opt in item.options]
            group_data[item.name] = options
        grouped_menu[group.name] = group_data

    return render_template("order.html",
                           current_user=current_user,
                           session=session,
                           grouped_menu=grouped_menu,
                           user_budget=team_budget,
                           remaining_budget=remaining_budget,
                           week_range=week_range_str,
                           form_data=form_data)

@app.route('/add_to_order', methods=['POST'])
@login_required
def add_to_order():
    form_data = request.form.to_dict()
    print("DEBUG form_data:", form_data)

    cleaned_form = {}

    from models import MenuOption, MenuItem, MenuGroup

    # Find Hyvee and Produce group IDs
    allowed_groups = ["Hyvee", "Produce"]
    allowed_group_ids = [
        g.id for g in MenuGroup.query.filter(MenuGroup.name.in_(allowed_groups)).all()
    ]

    # Build a lookup: meta_key → group_id
    option_group_lookup = {}
    for group in MenuGroup.query.all():
        for item in group.items:
            for opt in item.options:
                key = f"meta_{item.name.replace(' ', '_')}_{opt.name.replace(' ', '_')}"
                option_group_lookup[key] = group.id
    print("DEBUG option_group_lookup:", option_group_lookup)

    for key in form_data:
        if key.startswith("qty_"):
            qty_str = form_data[key].strip()
            if not qty_str:
                continue

            try:
                qty = float(qty_str)
            except ValueError:
                continue

            if qty <= 0:
                continue

            suffix = key[4:]
            meta_key = f"meta_{suffix}"
            group_id = option_group_lookup.get(meta_key)

            print(f"DEBUG: qty key={key}, meta_key={meta_key}, group_id={group_id}, qty={qty}")

            # Validate quantity based on group
            if group_id in allowed_group_ids:
                # Hyvee/Produce: allow .5 increments
                if qty % 0.5 != 0:
                    print(f"Invalid: {qty} for allowed group {group_id}")
                    continue
            else:
                # Other groups: only integers
                if not qty.is_integer():
                    print(f"Invalid: {qty} for non-allowed group {group_id}")
                    continue
                qty = int(qty)

            # Save validated qty and meta
            cleaned_form[f"qty_{suffix}"] = str(qty)
            cleaned_form[meta_key] = form_data.get(meta_key, "")

    session['last_order_form'] = cleaned_form
    print("DEBUG cleaned_form:", cleaned_form)

    if form_data.get("action") == "review":
        return redirect(url_for('review_order'))
    else:
        return redirect(url_for('submit_order'))

@app.route('/order/edit', methods=['POST'])
@login_required
def order_form_edit():
    if current_user.name == 'admin' and not session.get("admin_as_football"):
        return redirect(url_for('admin_dashboard'))

    form_data = request.form
    selected_items = []

    for key in form_data:
        if key.startswith("meta_"):
            suffix = key[5:]
            qty_key = f"qty_{suffix}"
            qty_str = form_data.get(qty_key, "").strip()

            try:
                quantity = float(qty_str)
                if quantity <= 0:
                    continue
            except ValueError:
                continue

            try:
                item_name, option_name, price = form_data[key].split("|||")
                selected_items.append({
                    "name": item_name,
                    "option": option_name,
                    "price": float(price),
                    "quantity": quantity,   # use float, not int
                    "meta_key": key,
                    "qty_key": qty_key
                })
            except Exception:
                continue

    return render_template("order_edit.html", selected_items=selected_items)

@app.route('/order/review', methods=['GET', 'POST'])
@login_required
def review_order():
    # Get team from current user
    team = current_user.team
    if not team:
        return "Team not found for current user.", 400

    team_budget = team.budget

    # Week range string
    today = datetime.now(CDT)
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)
    week_range_str = f"{start_of_week.strftime('%-m/%-d/%y')} - {end_of_week.strftime('%-m/%-d/%y')}"

    # Load last order form data from session
    form_data = session.get('last_order_form', {})
    session['last_order_form'] = form_data

    items = []
    total = 0.0

    for key in form_data:
        if key.startswith("meta_"):
            suffix = key[5:]
            qty_key = f"qty_{suffix}"
            qty_str = form_data.get(qty_key, "0").strip()

            try:
                quantity = float(qty_str)
            except ValueError:
                continue

            if quantity <= 0:
                continue

            try:
                item_name, option_name, price = form_data[key].split("|||")
                price = float(price)
                subtotal = round(price * quantity, 2)
                total += subtotal
                items.append({
                    "name": item_name,
                    "option": option_name,
                    "price": price,
                    "quantity": quantity,
                    "subtotal": subtotal
                })
            except Exception:
                continue

    return render_template("order_review.html",
                           items=items,
                           total=round(total, 2),
                           user_budget=team_budget,
                           remaining_budget=round(team.remaining_budget - total, 2),
                           week_range=week_range_str,
                           form_data=form_data)

@app.route('/order/submit', methods=['POST'])
@login_required
def finalize_order():
    form_data = session.get('last_order_form', {})
    items = []
    total = 0.0

    for key in form_data:
        if key.startswith("meta_"):
            suffix = key[5:]
            qty_key = f"qty_{suffix}"
            qty_str = form_data.get(qty_key, "0").strip()

            try:
                quantity = float(qty_str)
            except ValueError:
                continue

            if quantity <= 0:
                continue

            try:
                item_name, option_name, price = form_data[key].split("|||")
                price = float(price)
                total += price * quantity
                items.append({
                    "name": item_name,
                    "option": option_name,
                    "price": price,
                    "quantity": quantity
                })
            except:
                continue

    # ✅ Save to DB if there are items
    if items:
        now = datetime.now(CDT)
        new_order = Order(
            user_id=current_user.id,
            team_id=current_user.team_id,
            date=now.date(),
            time=now.time()
        )
        db.session.add(new_order)
        db.session.flush()  # Get new_order.id for OrderItem foreign keys

        for item in items:
            db.session.add(OrderItem(
                order_id=new_order.id,
                item_name=item["name"],
                option_name=item["option"],
                quantity=item["quantity"],
                price=item["price"]
            ))

            # ✅ SUBTRACT FROM INVENTORY
            menu_option = MenuOption.query.join(MenuItem).filter(
                MenuItem.name == item["name"],
                MenuOption.name == item["option"]
            ).first()

            if menu_option:
                menu_option.quantity = max(0, (menu_option.quantity or 0) - item["quantity"])

        # ✅ Adjust budget after all items
        team = current_user.team
        team.remaining_budget -= total

        print("==== FINALIZE ORDER: about to save ====")
        for item in items:
            print(f"{item['name']} - {item['option']} : {item['quantity']}")

        db.session.commit()

    # ✅ Clear form session after saving
    session.pop('last_order_form', None)

    return redirect(url_for('submit_order'))

@app.route('/admin')
@login_required
def admin_dashboard():
    # Access control: allow full admin or KSU Football limited admin
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    # Query all team names from the database
    all_teams = [team.name for team in Team.query.order_by(Team.name).all()]

    # Week range string
    today = datetime.now(CDT)
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)
    week_range_str = f"{start_of_week.strftime('%-m/%-d/%y')} - {end_of_week.strftime('%-m/%-d/%y')}"

    return render_template('admin_dashboard.html', teams=all_teams, week_range=week_range_str)

@app.route('/admin/dashboard')
@login_required
def admin_dashboard_view():
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    teams = [team.name for team in Team.query.order_by(Team.name).all()]

    return render_template("admin_dashboard.html", teams=teams)

@app.route('/admin/produce_hyvee')
@login_required
def admin_produce_hyvee():
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    # Load structured menu to get Produce + Hyvee items
    groups = MenuGroup.query.filter(MenuGroup.name.in_(["Produce", "Hyvee"])).all()
    valid_items = {item.name for group in groups for item in group.items}

    # Define current week range
    today = datetime.now(CDT)
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)

    # Query all relevant orders within the week
    matching_orders = []
    orders = Order.query.filter(Order.date >= start_of_week.date(), Order.date <= end_of_week.date()).all()

    for order in orders:
        team_name = order.team.name
        for item in order.items:
            if item.item_name in valid_items:
                matching_orders.append((
                    order.date.strftime("%Y-%m-%d"),
                    team_name,
                    item.item_name,
                    item.quantity
                ))

    week_range = f"{start_of_week.strftime('%-m/%-d/%y')} - {end_of_week.strftime('%-m/%-d/%y')}"

    return render_template("admin_produce_hyvee.html", orders=matching_orders, week_range=week_range)

@app.route('/admin/produce_hyvee/export')
@login_required
def export_produce_hyvee_excel():
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    # Define current week range
    today = datetime.now(CDT)
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)

    # Load valid Produce and Hyvee item names
    groups = MenuGroup.query.filter(MenuGroup.name.in_(["Produce", "Hyvee"])).all()
    valid_items = {item.name for group in groups for item in group.items}

    # Create Excel workbook
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Produce & Hyvee"
    ws_out.append(["Date", "Team", "Item", "Quantity"])

    # Query orders within week
    orders = Order.query.filter(Order.date >= start_of_week.date(), Order.date <= end_of_week.date()).all()

    for order in orders:
        team_name = order.team.name
        order_date_str = order.date.strftime("%Y-%m-%d")
        for item in order.items:
            if item.item_name in valid_items:
                ws_out.append([order_date_str, team_name, item.item_name, item.quantity])

    # Create downloadable Excel file
    output = BytesIO()
    wb_out.save(output)
    output.seek(0)

    filename = f"Produce_Hyvee_Orders_{start_of_week.strftime('%Y%m%d')}.xlsx"
    return send_file(output,
                     download_name=filename,
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/admin/weekly_summary')
@login_required
def admin_weekly_summary():
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403
    
    today = datetime.now(CDT)
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)
    week_range_str = f"{start_of_week.strftime('%-m/%-d/%y')} - {end_of_week.strftime('%-m/%-d/%y')}"

    all_orders = []

    orders = Order.query.filter(Order.date >= start_of_week.date(), Order.date <= end_of_week.date()).all()

    for order in orders:
        team_name = order.team.name
        for item in order.items:
            item_full = f"{item.item_name} - {item.option_name}".strip(" -")
            all_orders.append({
                "date": order.date.strftime("%-m/%-d/%y"),
                "team": team_name,
                "item": item_full,
                "quantity": item.quantity
            })

    return render_template(
        "weekly_summary.html",
        week_range=week_range_str,
        orders=all_orders,
        datetime=datetime,   # ✅ add this
        timedelta=timedelta
    )

@app.route('/admin/weekly_summary/export')
@login_required
def export_weekly_summary_excel():
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    today = datetime.now(CDT)
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)

    orders = Order.query.filter(Order.date >= start_of_week.date(), Order.date <= end_of_week.date()).all()

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Weekly Summary"
    ws_out.append(["Date", "Team", "Item", "Quantity"])

    for order in orders:
        team_name = order.team.name
        order_date = order.date.strftime("%Y-%m-%d")
        for item in order.items:
            item_full = f"{item.item_name} - {item.option_name}".strip(" -")
            ws_out.append([order_date, team_name, item_full, item.quantity])

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)

    filename = f"Full_Weekly_Orders_{start_of_week.strftime('%Y%m%d')}.xlsx"
    return send_file(output,
                     download_name=filename,
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/admin/football_order')
@login_required
def admin_football_order():
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    # Only full admin simulates KSU Football ordering (without logging in as them)
    session['admin_as_football'] = True
    return redirect(url_for('submit_order'))

@app.route('/admin/team/<team_name>')
@login_required
def view_team_orders(team_name):
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    team = Team.query.filter_by(name=team_name).first()
    if not team:
        return f"Team '{team_name}' not found", 404

    today = datetime.now(CDT)
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)
    week_range_str = f"{start_of_week.strftime('%-m/%-d/%y')} - {end_of_week.strftime('%-m/%-d/%y')}"

        # Build price lookup from DB
    price_lookup = {}
    all_items = MenuItem.query.all()
    for item in all_items:
        for opt in item.options:
            key = f"{item.name}|||{opt.name}"
            price_lookup[key] = opt.price

    # Init structures
    weekly_orders_by_member = {}
    all_totals = {}
    total_cost = 0.0

    for user in team.members:
        member_orders = []
        member_total = 0.0

        for order in user.orders:
            for item in order.items:
                key = f"{item.item_name}|||{item.option_name}"
                price = price_lookup.get(key, 0.0)
                quantity = item.quantity or 0
                subtotal = quantity * price

                # Weekly filter
                if start_of_week.date() <= order.date <= end_of_week.date():
                    member_orders.append({
                        "date": order.date.strftime("%Y-%m-%d"),
                        "time": order.time.strftime("%I:%M %p"),
                        "item": f"{item.item_name} - {item.option_name}".strip(" -"),
                        "quantity": quantity,
                        "price": f"${price:.2f}",
                        "subtotal": f"${subtotal:.2f}"
                    })
                    member_total += subtotal
                    total_cost += subtotal

                # Accumulate full-year totals
                full_item_name = f"{item.item_name} - {item.option_name}".strip(" -")
                if full_item_name not in all_totals:
                    all_totals[full_item_name] = {"qty": 0, "total_cost": 0.0}

                all_totals[full_item_name]["qty"] += quantity
                all_totals[full_item_name]["total_cost"] += quantity * price

        if member_orders:
            weekly_orders_by_member[user.name] = {
                "orders": member_orders,
                "total": member_total
            }

    # Team budget info
    team_budget = team.budget
    remaining_budget = round(team.remaining_budget, 2)

    return render_template("team_orders.html",
                           team_name=team_name,
                           week_range=week_range_str,
                           weekly_orders_by_member=weekly_orders_by_member,
                           total_orders=all_totals,
                           total_cost=total_cost,
                           user_budget=team_budget,
                           remaining_budget=remaining_budget)

@app.route('/admin/user/<user_name>')
@login_required
def view_user_file(user_name):
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    # Find the user by name
    user = User.query.filter_by(name=user_name).first()
    if not user:
        return f"No user found with name '{user_name}'", 404

    current_week = datetime.now(CDT).isocalendar()[1]
    weekly_orders = []
    item_totals = {}

    for order in user.orders:
        order_week = order.date.isocalendar()[1]

        for item in order.items:
            item_key = f"{item.item_name} - {item.option_name}".strip(" -")

            # Running total of all orders (lifetime)
            if item_key not in item_totals:
                item_totals[item_key] = 0
            item_totals[item_key] += item.quantity

            # Weekly orders (current week only)
            if order_week == current_week:
                weekly_orders.append({
                    "date": order.date.strftime("%Y-%m-%d"),
                    "item": item_key,
                    "quantity": item.quantity
                })

    total_orders = [{"item": name, "quantity": qty} for name, qty in item_totals.items()]

    return render_template("user_orders.html",
                           user_name=user_name,
                           weekly_orders=weekly_orders,
                           total_orders=total_orders)

@app.route('/admin/weekly_totals')
@login_required
def weekly_totals():
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    # Build price lookup from the database
    price_lookup = {
        f"{item.name}|||{opt.name}": opt.price
        for item in MenuItem.query.all()
        for opt in item.options
    }

    yearly_totals_by_week = {}
    all_years = set()

    orders = Order.query.join(Team).all()
    for order in orders:
        team_name = order.team.name
        year = order.date.year
        week = get_week_number(order.date)

        all_years.add(year)

        if week < 1 or week > 52:
            continue

        yearly_totals_by_week.setdefault(year, {}).setdefault(week, {})

        for item in order.items:
            key = f"{item.item_name}|||{item.option_name}"
            price = price_lookup.get(key, 0.0)
            subtotal = item.quantity * price
            current = yearly_totals_by_week[year][week].get(team_name, 0.0)
            yearly_totals_by_week[year][week][team_name] = current + subtotal

    # Ensure every week/year has all teams
    all_team_names = [team.name for team in Team.query.order_by(Team.name).all()]
    for year in all_years:
        for week in range(1, 53):
            yearly_totals_by_week.setdefault(year, {}).setdefault(week, {})
            for team in all_team_names:
                yearly_totals_by_week[year][week].setdefault(team, 0.0)

    return render_template(
        "weekly_totals.html",
        yearly_totals_by_week=yearly_totals_by_week,
        users=all_team_names,
        years=sorted(all_years),
        datetime=datetime,
        timedelta=timedelta
    )

@app.route('/download_weekly_totals')
@login_required
def download_weekly_totals():
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    # Build price lookup
    price_lookup = {
        f"{item.name}|||{opt.name}": opt.price
        for item in MenuItem.query.all()
        for opt in item.options
    }

    # Build yearly totals by week
    yearly_totals_by_week = {}
    all_years = set()
    all_teams = [team.name for team in Team.query.order_by(Team.name).all()]

    orders = Order.query.join(Team).all()
    for order in orders:
        team_name = order.team.name
        year = order.date.year
        week = get_week_number(order.date)

        all_years.add(year)
        if week < 1 or week > 52:
            continue

        yearly_totals_by_week.setdefault(year, {}).setdefault(week, {})
        for item in order.items:
            key = f"{item.item_name}|||{item.option_name}"
            price = price_lookup.get(key, 0.0)
            subtotal = item.quantity * price
            current = yearly_totals_by_week[year][week].get(team_name, 0.0)
            yearly_totals_by_week[year][week][team_name] = current + subtotal

    # Fill missing values with 0 for all teams
    for year in all_years:
        for week in range(1, 53):
            yearly_totals_by_week.setdefault(year, {}).setdefault(week, {})
            for team in all_teams:
                yearly_totals_by_week[year][week].setdefault(team, 0.0)

    # Write to Excel
    from openpyxl import Workbook
    from io import BytesIO
    from datetime import datetime, timedelta

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Totals"

    year = max(all_years)
    july_first = datetime(year, 7, 1).date()
    if july_first.weekday() != 6:
        july_first = july_first - timedelta(days=july_first.weekday() + 1)

    headers = ["Week #", "Date"] + all_teams
    ws.append(headers)

    for week in range(1, 53):
        start_of_week = july_first + timedelta(weeks=week - 1)
        end_of_week = start_of_week + timedelta(days=6)
        row = [f"Week {week}", f"{start_of_week.strftime('%-m/%-d/%y')} - {end_of_week.strftime('%-m/%-d/%y')}"]
        for team in all_teams:
            value = round(yearly_totals_by_week[year][week][team], 2)
            row.append(value)
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     download_name="weekly_totals.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/admin/all_orders')
@login_required
def all_orders():
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    all_orders = []

    all_db_orders = Order.query.join(User).join(Team).all()

    for order in all_db_orders:
        order_date = order.date
        order_time = order.time
        member_name = order.user.name
        team_name = order.team.name
        year = order_date.year
        week_num = get_week_number(order_date)

        for item in order.items:
            all_orders.append({
                "date": order_date.strftime("%Y-%m-%d"),
                "time": order_time.strftime("%I:%M %p"),
                "week": week_num,
                "year": year,
                "team": team_name,
                "member": member_name,
                "item": f"{item.item_name} - {item.option_name}".strip(" -"),
                "quantity": item.quantity
            })

    return render_template("all_orders.html", orders=all_orders)

@app.route('/admin/inventory', methods=['GET', 'POST'])
@login_required
def view_inventory():
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    if request.method == 'POST':
        for key, value in request.form.items():
            if key.startswith('quantity_'):
                option_id = key.replace('quantity_', '')
                option = MenuOption.query.get(int(option_id))
                if option:
                    try:
                        option.quantity = int(value)
                    except ValueError:
                        continue
        db.session.commit()
        return redirect(url_for('view_inventory'))

    # === Grouped menu view ===
    grouped_menu = OrderedDict()
    groups = MenuGroup.query.order_by(MenuGroup.position).all()
    for group in groups:
        items = MenuItem.query.filter_by(group_id=group.id).order_by(MenuItem.position).all()
        for item in items:
            item.options_data = MenuOption.query.filter_by(item_id=item.id).order_by(MenuOption.position).all()
        grouped_menu[group.name] = items

    # === Reorder Items (where quantity <= reorder point) ===
    reorder_items = MenuOption.query\
        .join(MenuItem, MenuOption.item_id == MenuItem.id)\
        .join(MenuGroup, MenuItem.group_id == MenuGroup.id)\
        .filter(MenuOption.quantity <= MenuOption.reorder_point)\
        .order_by(MenuGroup.name, MenuItem.name)\
        .all()

    return render_template('inventory.html', grouped_menu=grouped_menu, reorder_items=reorder_items)

@app.route('/admin/budgets', methods=['GET', 'POST'])
@login_required
def manage_budgets():
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    teams = Team.query.order_by(Team.name).all()

    if request.method == 'POST':
        if 'reset' in request.form:
            # Reset remaining budgets
            for team in teams:
                team.remaining_budget = team.budget
            db.session.commit()
            return render_template("manage_budgets.html", team_budgets={team.name: team.budget for team in teams}, message="✅ All remaining budgets reset.")
        else:
            # Update budgets
            for team in teams:
                new_value = request.form.get(team.name)
                if new_value:
                    try:
                        team.budget = float(new_value)
                    except ValueError:
                        continue
            db.session.commit()
            return redirect(url_for('manage_budgets'))

    team_budgets = {team.name: team.budget for team in teams}
    return render_template("manage_budgets.html", team_budgets=team_budgets)

@app.route('/admin/edit_menu', methods=['GET', 'POST'])
@login_required
def edit_menu():
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    if request.method == 'POST':
        form = request.form
        rename_map = {}

        # Build mapping: old group key → new group name
        for key in form:
            if key.startswith("group_rename["):
                original_name = key.split("group_rename[")[1].split("]")[0]
                new_name = form[key].strip()
                if new_name:
                    rename_map[original_name] = new_name

        try:
            for group_key, group_name in rename_map.items():
                group = MenuGroup.query.filter_by(name=group_name).first()
                if not group:
                    last_group = MenuGroup.query.order_by(MenuGroup.position.desc()).first()
                    next_position = (last_group.position + 1) if last_group else 1
                    group = MenuGroup(name=group_name, position=next_position)
                    db.session.add(group)
                    db.session.flush()

                # Load item names for this group
                item_names = form.getlist(f'group_names[{group_key}][item_names][]')
                submitted_item_names = set(name.strip() for name in item_names if name.strip())

                for item_name in submitted_item_names:
                    item = MenuItem.query.filter_by(name=item_name, group_id=group.id).first()
                    if not item:
                        last_item = MenuItem.query.filter_by(group_id=group.id).order_by(MenuItem.position.desc()).first()
                        next_item_pos = (last_item.position + 1) if last_item else 1
                        item = MenuItem(name=item_name, group_id=group.id, position=next_item_pos)
                        db.session.add(item)
                        db.session.flush()

                    # Load options and prices
                    options = form.getlist(f'options[{item_name}][]')
                    prices = form.getlist(f'prices[{item_name}][]')

                    if not options or not prices or len(options) != len(prices):
                        continue

                    existing_options = {opt.name: opt for opt in item.options}
                    updated_option_names = set()

                    for opt_name, price_str in zip(options, prices):
                        opt_name = opt_name.strip()
                        updated_option_names.add(opt_name)

                        try:
                            price = float(price_str)
                        except ValueError:
                            continue

                        option = existing_options.get(opt_name)
                        if not option:
                            option = MenuOption(name=opt_name, item_id=item.id)
                            db.session.add(option)

                        option.price = price

                    # Delete removed options if quantity is 0
                    for opt_name, opt in existing_options.items():
                        if opt_name not in updated_option_names:
                            if (opt.quantity or 0) == 0:
                                db.session.delete(opt)

                # Delete removed items from this group
                existing_items = MenuItem.query.filter_by(group_id=group.id).all()
                for existing_item in existing_items:
                    if existing_item.name not in submitted_item_names:
                        for opt in existing_item.options:
                            db.session.delete(opt)
                        db.session.delete(existing_item)

            # Delete removed groups
            all_existing_groups = MenuGroup.query.all()
            submitted_group_names = set(rename_map.values())

            for group in all_existing_groups:
                if group.name not in submitted_group_names:
                    items = MenuItem.query.filter_by(group_id=group.id).all()
                    for item in items:
                        for opt in item.options:
                            db.session.delete(opt)
                        db.session.delete(item)
                    db.session.delete(group)

            db.session.commit()
            return redirect(url_for('edit_menu'))

        except Exception as e:
            db.session.rollback()
            raise

    # === GET: Load current grouped menu ===
    grouped_menu = OrderedDict()
    groups = MenuGroup.query.order_by(MenuGroup.position).all()
    for group in groups:
        group_data = OrderedDict()
        items = MenuItem.query.filter_by(group_id=group.id).order_by(MenuItem.position).all()
        for item in items:
            options = [
                {"name": opt.name, "price": opt.price}
                for opt in MenuOption.query.filter_by(item_id=item.id).order_by(MenuOption.position).all()
            ]
            group_data[item.name] = options
        grouped_menu[group.name] = group_data

    return render_template('edit_menu_fixed.html', grouped_menu=grouped_menu)

@app.route('/admin/edit_users', methods=['GET', 'POST'])
@login_required
def edit_users():
    if not (session.get('admin_as_football') or session.get('member_name') == "Scott Trausch"):
        return "Access Denied", 403

    if request.method == 'POST':
        team_names = request.form.getlist('team_names[]')
        members_list = request.form.getlist('members[]')

        # ✅ Keep track of team names submitted in the form
        updated_team_names = set()

        for team_name, members_raw in zip(team_names, members_list):
            team_name = team_name.strip()
            if not team_name:
                continue

            updated_team_names.add(team_name)

            # ✅ Find or create the team
            team = Team.query.filter_by(name=team_name).first()
            if not team:
                team = Team(name=team_name, budget=100.0)
                db.session.add(team)
                db.session.flush()

            # ✅ Deactivate all current users on this team
            for existing_user in User.query.filter_by(team_id=team.id).all():
                existing_user.is_enabled = False

            # ✅ Reactivate or create listed users
            members = [m.strip() for m in members_raw.splitlines() if m.strip()]
            for member_name in members:
                existing = User.query.filter_by(name=member_name, team_id=team.id).first()
                if existing:
                    existing.is_enabled = True
                else:
                    db.session.add(User(name=member_name, team_id=team.id, is_enabled=True))

        # ✅ Delete any teams not included in the current form submission
        all_teams = Team.query.all()
        for team in all_teams:
            if team.name not in updated_team_names:
                db.session.delete(team)  # ✅ Triggers team deletion; users' team_id will become NULL

        db.session.commit()
        return redirect(url_for('admin_dashboard'))

    # GET: Load current user assignments for display
    teams = Team.query.order_by(Team.name).all()
    users_by_team = OrderedDict()
    for team in teams:
        users_by_team[team.name] = [user.name for user in team.members if user.is_enabled]

    return render_template("edit_users.html", users=users_by_team)

@app.route('/admin/edit_inventory', methods=['GET', 'POST'])
@login_required
def edit_inventory():
    if request.method == 'POST':
        options = MenuOption.query.all()
        for option in options:
            cs_key = f"case_size_{option.id}"
            rp_key = f"reorder_point_{option.id}"

            case_size = request.form.get(cs_key, "").strip()
            reorder_point = request.form.get(rp_key, "").strip()

            # ✅ Save only if valid
            if case_size.isdigit():
                option.case_size = int(case_size)
            if reorder_point.isdigit():
                option.reorder_point = int(reorder_point)

        db.session.commit()
        return redirect(url_for('edit_inventory'))

    # GET view
    grouped_menu = OrderedDict()
    groups = MenuGroup.query.order_by(MenuGroup.position).all()
    for group in groups:
        items = MenuItem.query.filter_by(group_id=group.id).order_by(MenuItem.position).all()
        for item in items:
            item.options_data = MenuOption.query.filter_by(item_id=item.id).order_by(MenuOption.position).all()
        grouped_menu[group.name] = items

    return render_template('edit_inventory.html', grouped_menu=grouped_menu)

def start_scheduler():
    scheduler = BackgroundScheduler(timezone=CDT)
    scheduler.add_job(
        func=email_all_orders,
        trigger='cron',
        day_of_week='sun',
        hour=23,
        minute=59,
        id='weekly_all_orders_email'
    )
    scheduler.start()
    print("✅ Scheduler started.")

with app.app_context():
    start_scheduler()

@app.route('/admin/test_email')
@login_required
def test_email():
    from tasks import email_all_orders
    email_all_orders()
    return "✅ Test email sent (check logs for success/failure)."

#@app.route('/admin/clear_orders')
#@login_required
#def clear_orders():
    #if session.get('member_name') != "Scott Trausch":
        #return "Access Denied", 403

    #OrderItem.query.delete()
    #Order.query.delete()
    #db.session.commit()
    #return "✅ All orders cleared."

# === Run the App ===
if __name__ == '__main__':
    app.run(debug=True)
